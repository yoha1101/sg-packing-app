import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict, defaultdict
import io
import re
from datetime import datetime

st.set_page_config(page_title="SG Export Document Generator", layout="centered")
st.title("📦 SPECIALGUEST Export Document Generator")

CBM_PER_BOX = 0.088

PRIORITY = [
    'Insulated Snowboard Jacket', 'Snowboard Jacket', 'Snowboard Pants',
    'Long-sleeve Tee', 'Sweatshirt, Hooded Sweatshirt', 'Sweatpants', 'T-shirt',
    'Cap', 'Beanie', 'Beanie(Angora)', 'HOOD WARMER N', 'BALACLAVA', 'Snowboard Stomppad'
]

SHEET_ORDER = ['Snowboard Pants','Snowboard Jacket','Insulated Snowboard Jacket',
    'Sweatpants','Sweatshirt, Hooded Sweatshirt','Long-sleeve Tee','T-shirt',
    'Beanie','Beanie(Angora)','Cap','HOOD WARMER N','BALACLAVA','Snowboard Stomppad']

MAIN_COLORS = {
    'Snowboard Pants':'E2F0D9','Snowboard Jacket':'FFF2CC',
    'Insulated Snowboard Jacket':'FCE4D6','Sweatpants':'EAD1DC',
    'Sweatshirt, Hooded Sweatshirt':'D9E1F2','Long-sleeve Tee':'DDEBF7',
    'T-shirt':'F2F2F2','Beanie':'E2EFDA','Beanie(Angora)':'FFF2CC',
    'Cap':'DAEEF3','HOOD WARMER N':'F4CCCC','BALACLAVA':'D9EAD3',
    'Snowboard Stomppad':'D9D9D9',
}
SEC_COLORS = {
    'Snowboard Pants':'C6EFCE','Snowboard Jacket':'FFEB9C',
    'Insulated Snowboard Jacket':'F8CBAD','Sweatpants':'D5A6BD',
    'Sweatshirt, Hooded Sweatshirt':'B4C7E7','Long-sleeve Tee':'BDD7EE',
    'T-shirt':'D9D9D9','Beanie':'C6D9B8','Beanie(Angora)':'FFE699',
    'Cap':'B8D9E8','HOOD WARMER N':'EA9999','BALACLAVA':'A9C9A4',
    'Snowboard Stomppad':'BFBFBF',
}

def get_prio(d):
    return PRIORITY.index(d) if d in PRIORITY else 99

def standardize_category(style_raw):
    style = str(style_raw).strip().lower()
    if 'snowboard pants' in style or 'cargo pants' in style: return 'Snowboard Pants'
    elif 'snowboard jacket' in style and 'insulated' not in style: return 'Snowboard Jacket'
    elif 'insulated' in style and 'jacket' in style: return 'Insulated Snowboard Jacket'
    elif 'sweatpants' in style: return 'Sweatpants'
    elif 'sweatshirt' in style or 'hooded' in style or 'hoodie' in style: return 'Sweatshirt, Hooded Sweatshirt'
    elif 'long-sleeve' in style or 'long sleeve' in style: return 'Long-sleeve Tee'
    elif 't-shirt' in style or 'tee' in style: return 'T-shirt'
    elif 'cap' in style: return 'Cap'
    elif 'angora' in style: return 'Beanie(Angora)'
    elif 'beanie' in style: return 'Beanie'
    elif 'hood warmer' in style: return 'HOOD WARMER N'
    elif 'balaclava' in style: return 'BALACLAVA'
    elif 'stomppad' in style or 'stomp pad' in style: return 'Snowboard Stomppad'
    return style_raw

thin = Side(style='thin')
def tb(): return Border(left=thin, right=thin, top=thin, bottom=thin)
center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_a  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
hdr_fill = PatternFill('solid', fgColor='006FC0')
col_fill = PatternFill('solid', fgColor='D9D9D9')
alt_fill = PatternFill('solid', fgColor='F7F6D6')
no_fill  = PatternFill()

def sc(ws, row, col, val, bold=False, size=9, color='000000', fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, size=size, color=color)
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c

def parse_export_csv(csv_file):
    df = pd.read_csv(csv_file)
    df['박스번호'] = df['박스번호'].ffill()
    df['무게(kg)'] = df['무게(kg)'].ffill()
    df['Category'] = df['Style'].apply(standardize_category)
    boxes = {}
    for box_no, group in df.groupby('박스번호'):
        boxes[box_no] = {'box_no': box_no, 'weight': group['무게(kg)'].iloc[0], 'items': []}
        for _, row in group.iterrows():
            boxes[box_no]['items'].append({
                'sku': row['SKU'], 'style': row['Style'], 'category': row['Category'],
                'item_name': row['품목명'], 'hs_code': row['HS Code'], 'color': row['Color'],
                'size': row['Size'], 'qty': row['수량'], 'price': row['단가(KRW)'],
                'material': row['Material']
            })
    return df, boxes

def make_packing_list(boxes, destination):
    wb = Workbook(); ws = wb.active; ws.title = destination
    ws.column_dimensions['A'].width = 9; ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 20
    for col in ['D','E','F','G','H','I']: ws.column_dimensions[col].width = 6
    ws.column_dimensions['J'].width = 8; ws.column_dimensions['K'].width = 9
    headers = ['CTN NO.','STYLE','COLOR','S','M','L','XL','2XL','3XL','TOTAL','G.W(kgs)']
    for i, h in enumerate(headers, 1): sc(ws, 1, i, h, bold=True, fill=col_fill, align=center, border=tb())
    cur_row = 2
    for box_no in sorted(boxes.keys()):
        box = boxes[box_no]
        style_data = defaultdict(lambda: {'color':'','sizes':{'S':0,'M':0,'L':0,'XL':0,'2XL':0,'3XL':0}})
        for item in box['items']:
            key = item['item_name']; style_data[key]['color'] = item['color']
            size = item['size']
            if size in style_data[key]['sizes']: style_data[key]['sizes'][size] += item['qty']
        first_in_box = True; box_first_row = cur_row
        for style, data in style_data.items():
            if first_in_box: sc(ws, cur_row, 1, box_no, bold=True, align=center, border=tb()); first_in_box = False
            else: sc(ws, cur_row, 1, '', border=tb())
            sc(ws, cur_row, 2, style, align=left_a, border=tb())
            sc(ws, cur_row, 3, data['color'], align=left_a, border=tb())
            total = 0
            for i, size in enumerate(['S','M','L','XL','2XL','3XL'], 4):
                qty = data['sizes'][size]; sc(ws, cur_row, i, qty if qty > 0 else '', align=center, border=tb()); total += qty
            sc(ws, cur_row, 10, total, align=center, border=tb())
            sc(ws, cur_row, 11, box['weight'] if cur_row == box_first_row else '', align=center, border=tb())
            cur_row += 1
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def make_invoice(df, messrs, destination, date_str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_right_only = Border(right=thin)
    
    font_bold_aptos_9 = Font(name='Aptos', bold=True, size=9)
    font_normal_aptos_8 = Font(name='Aptos', size=8)
    
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right_wrap = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    cell = ws.cell(row=1, column=1, value='NO.')
    cell.font = font_bold_aptos_9
    cell.alignment = align_center_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=2, value='HS CODE')
    cell.font = font_bold_aptos_9
    cell.alignment = align_center_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=3, value='Description of goods')
    cell.font = font_normal_aptos_8
    cell.alignment = align_left_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=4)
    cell.border = border_right_only
    ws.merge_cells('C1:D1')
    
    cell = ws.cell(row=1, column=5, value='Fabric ratio')
    cell.font = font_normal_aptos_8
    cell.alignment = align_left_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=6)
    cell.border = border_right_only
    
    cell = ws.cell(row=1, column=7)
    cell.border = border_right_only
    ws.merge_cells('E1:G1')
    
    cell = ws.cell(row=1, column=8, value='QTY')
    cell.font = font_normal_aptos_8
    cell.alignment = align_right_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=9, value='Unit Price(KRW)')
    cell.font = font_normal_aptos_8
    cell.alignment = align_right_wrap
    cell.border = border_all
    
    cell = ws.cell(row=1, column=10, value='Amount(KRW)')
    cell.font = font_normal_aptos_8
    cell.alignment = align_right_wrap
    cell.border = border_all
    
    grouped = df.groupby(['품목명', 'HS Code', 'Material', '단가(KRW)']).agg({'수량': 'sum'}).reset_index()
    
    cur_row = 2
    total_qty = 0
    total_amount = 0
    
    for idx, (_, row) in enumerate(grouped.iterrows(), 1):
        qty = int(row['수량'])
        unit_price = int(row['단가(KRW)'])
        amount = qty * unit_price
        
        cell_a = ws.cell(row=cur_row, column=1, value=idx)
        cell_a.font = font_bold_aptos_9
        cell_a.alignment = align_left_wrap
        cell_a.border = border_all
        
        cell_b = ws.cell(row=cur_row, column=2, value=row['HS Code'])
        cell_b.font = font_bold_aptos_9
        cell_b.alignment = align_left_wrap
        cell_b.border = border_all
        
        cell_c = ws.cell(row=cur_row, column=3, value=row['품목명'])
        cell_c.font = font_normal_aptos_8
        cell_c.alignment = align_left_wrap
        cell_c.border = border_all
        
        cell_d = ws.cell(row=cur_row, column=4)
        cell_d.border = border_right_only
        
        ws.merge_cells(f'C{cur_row}:D{cur_row}')
        
        cell_e = ws.cell(row=cur_row, column=5, value=row['Material'])
        cell_e.font = font_normal_aptos_8
        cell_e.alignment = align_left_wrap
        cell_e.border = border_all
        
        cell_f = ws.cell(row=cur_row, column=6)
        cell_f.border = border_right_only
        
        cell_g = ws.cell(row=cur_row, column=7)
        cell_g.border = border_right_only
        
        ws.merge_cells(f'E{cur_row}:G{cur_row}')
        
        cell_h = ws.cell(row=cur_row, column=8, value=qty)
        cell_h.font = font_normal_aptos_8
        cell_h.alignment = align_right_wrap
        cell_h.border = border_all
        
        cell_i = ws.cell(row=cur_row, column=9, value=unit_price)
        cell_i.font = font_normal_aptos_8
        cell_i.alignment = align_right_wrap
        cell_i.border = border_all
        
        cell_j = ws.cell(row=cur_row, column=10, value=amount)
        cell_j.font = font_normal_aptos_8
        cell_j.alignment = align_right_wrap
        cell_j.border = border_all
        
        total_qty += qty
        total_amount += amount
        cur_row += 1
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def make_actual_packing_list(boxes, messrs, destination, date_str):
    wb = Workbook(); ws = wb.active; ws.title = 'ACTUAL PACKING LIST'
    widths = {'A':16,'B':34,'C':30,'D':9,'E':52,'F':14,'G':11,'H':9}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    ws.merge_cells('A1:H1'); sc(ws, 1, 1, 'ACTUAL PACKING LIST', bold=True, size=12, color='FFFFFF', fill=hdr_fill, align=center)
    ws.row_dimensions[1].height = 20; sc(ws, 2, 1, 'SPECIALGUEST®', bold=True, size=10)
    ws.merge_cells('A3:H3'); sc(ws, 3, 1, 'postal code: 12923 / 22, Misagangbyeonhangang-ro 346beon-gil, Hanam-si, Gyeonggi-do, Republic of Korea', align=left_a)
    ws.merge_cells('A4:H4'); sc(ws, 4, 1, 'Tel: +82 7077643333  Email: specialguest.co.kr@gmail.com', align=left_a)
    ws.merge_cells('A6:F6'); sc(ws, 6, 1, f'MESSRS: {messrs}', bold=True)
    ws.merge_cells('G6:H6'); sc(ws, 6, 7, f'DATE: {date_str}', size=8)
    ws.merge_cells('A7:F7'); sc(ws, 7, 1, 'SHIPMENT FROM: Republic of Korea', bold=True)
    ws.merge_cells('A8:H8'); sc(ws, 8, 1, f'FINAL DESTINATION: {destination}', bold=True)
    sc(ws, 9, 8, CBM_PER_BOX, bold=True, align=center)
    headers = ['HS code','Description of goods','Fabric ratio','Quantity','** Comments','Dimension (cm)','Weight(kg)','CBM']
    for i, h in enumerate(headers, 1): sc(ws, 10, i, h, bold=True, fill=col_fill, align=center, border=tb())
    ws.row_dimensions[10].height = 14
    category_data = defaultdict(lambda: {'hs_code':'','material':'','qty':0,'boxes':[],'weight':0})
    for box_no, box in boxes.items():
        for item in box['items']:
            cat = item['category']; category_data[cat]['hs_code'] = item['hs_code']
            category_data[cat]['material'] = item['material']; category_data[cat]['qty'] += item['qty']
            if box_no not in category_data[cat]['boxes']:
                category_data[cat]['boxes'].append(box_no); category_data[cat]['weight'] += box['weight']
    sorted_cats = sorted(category_data.keys(), key=get_prio)
    cur_row = 11; use_alt = False; total_qty = 0; total_boxes = 0; total_weight = 0
    for cat in sorted_cats:
        data = category_data[cat]; fill = alt_fill if use_alt else no_fill; use_alt = not use_alt
        num_boxes = len(data['boxes'])
        sc(ws, cur_row, 1, data['hs_code'], bold=True, fill=fill, align=center, border=tb())
        sc(ws, cur_row, 2, cat, bold=True, fill=fill, align=left_a, border=tb())
        sc(ws, cur_row, 3, data['material'], bold=True, fill=fill, align=left_a, border=tb())
        sc(ws, cur_row, 4, data['qty'], bold=True, fill=fill, align=center, border=tb())
        sc(ws, cur_row, 5, num_boxes, bold=True, fill=fill, align=center, border=tb())
        sc(ws, cur_row, 6, '55 x 40 x 40', bold=True, fill=fill, align=center, border=tb())
        sc(ws, cur_row, 7, data['weight'], bold=True, fill=fill, align=center, border=tb())
        cbm_cell = ws.cell(row=cur_row, column=8, value=f'=E{cur_row}*$H$9')
        cbm_cell.font = Font(name='Arial', bold=True, size=9); cbm_cell.fill = fill
        cbm_cell.alignment = center; cbm_cell.border = tb()
        total_qty += data['qty']; total_boxes += num_boxes; total_weight += data['weight']; cur_row += 1
    cur_row += 1; sc(ws, cur_row, 1, 'Please check the ** comments.', bold=True); cur_row += 1
    sc(ws, cur_row, 1, 'N = Nylon / P = Polyester / PU = Polyurethane / C = Cotton / R = Rayon / A = Acrylic', bold=True)
    cur_row += 2
    for i, h in enumerate(['Quantity','CTNS','Dimension (cm)','Weight(kg)','CBM'], start=4):
        sc(ws, cur_row, i, h, bold=True, align=center)
    cur_row += 1; ws.merge_cells(f'A{cur_row}:C{cur_row}'); sc(ws, cur_row, 1, 'Total', bold=True)
    sc(ws, cur_row, 4, total_qty, bold=True, align=center)
    sc(ws, cur_row, 5, total_boxes, bold=True, align=center)
    sc(ws, cur_row, 6, '55 x 40 x 40', bold=True, align=center)
    sc(ws, cur_row, 7, total_weight, bold=True, align=center)
    sc(ws, cur_row, 8, round(total_boxes * CBM_PER_BOX, 3), bold=True, align=center)
    cur_row += 2; ws.merge_cells(f'D{cur_row}:H{cur_row}'); sc(ws, cur_row, 4, 'MADE IN KOREA', bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def make_category_packing_list(boxes):
    wb = Workbook(); wb.remove(wb.active)
    box_main_category = {}
    for box_no, box in boxes.items():
        cat_qty = defaultdict(int)
        for item in box['items']: cat_qty[item['category']] += item['qty']
        main_cat = max(cat_qty.items(), key=lambda x: (x[1], -get_prio(x[0])))[0]
        box_main_category[box_no] = main_cat
    category_boxes = defaultdict(list)
    for box_no, main_cat in box_main_category.items(): category_boxes[main_cat].append(box_no)
    for cat in SHEET_ORDER:
        if cat not in category_boxes: continue
        ws = wb.create_sheet(cat[:31]); main_fill = PatternFill('solid', fgColor=MAIN_COLORS.get(cat,'FFFFFF'))
        ws.column_dimensions['A'].width = 9; ws.column_dimensions['B'].width = 38; ws.column_dimensions['C'].width = 20
        for c in ['D','E','F','G','H','I']: ws.column_dimensions[c].width = 6
        ws.column_dimensions['J'].width = 8; ws.column_dimensions['K'].width = 9
        headers = ['CTN NO.','STYLE','COLOR','S','M','L','XL','2XL','3XL','TOTAL','G.W(kgs)']
        for i, h in enumerate(headers, 1): sc(ws, 1, i, h, bold=True, fill=col_fill, align=center, border=tb())
        cur_row = 2; total_qty = 0
        for box_no in sorted(category_boxes[cat]):
            box = boxes[box_no]
            style_data = defaultdict(lambda: {'color':'','category':'','sizes':{'S':0,'M':0,'L':0,'XL':0,'2XL':0,'3XL':0}})
            for item in box['items']:
                key = item['item_name']; style_data[key]['color'] = item['color']
                style_data[key]['category'] = item['category']; size = item['size']
                if size in style_data[key]['sizes']: style_data[key]['sizes'][size] += item['qty']
            first_in_box = True
            for style, data in style_data.items():
                row_fill = main_fill if data['category'] == cat else PatternFill('solid', fgColor=SEC_COLORS.get(data['category'],'EEEEEE'))
                if data['category'] == cat: total_qty += sum(data['sizes'].values())
                if first_in_box: sc(ws, cur_row, 1, box_no, bold=True, fill=row_fill, align=center, border=tb()); first_in_box = False
                else: sc(ws, cur_row, 1, '', fill=row_fill, border=tb())
                sc(ws, cur_row, 2, style, fill=row_fill, align=left_a, border=tb())
                sc(ws, cur_row, 3, data['color'], fill=row_fill, align=left_a, border=tb())
                row_total = 0
                for i, size in enumerate(['S','M','L','XL','2XL','3XL'], 4):
                    qty = data['sizes'][size]; sc(ws, cur_row, i, qty if qty > 0 else '', fill=row_fill, align=center, border=tb()); row_total += qty
                sc(ws, cur_row, 10, row_total, fill=row_fill, align=center, border=tb())
                sc(ws, cur_row, 11, box['weight'] if first_in_box else '', fill=row_fill, align=center, border=tb())
                cur_row += 1
        cur_row += 1; ws.merge_cells(f'A{cur_row}:B{cur_row}')
        sc(ws, cur_row, 1, 'TOTAL', bold=True, fill=col_fill); sc(ws, cur_row, 10, total_qty, bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

SIZE_COLS_SHEET = ['S','M','L','XL','2XL','3XL']
SIZE_ORDER_LIST = ['XS','S','M','L','XL','2XL','3XL','4XL','FREE']

def parse_sku(sku: str):
    for sz in ['3XL','2XL','XL','FREE','XS','S','M','L']:
        if str(sku).upper().endswith(sz): return str(sku)[:-len(sz)], sz
    return str(sku), ''

def get_season_from_sku(sku: str) -> str:
    m = re.search(r'\d{3}[WwSs](\d{2})', str(sku))
    if m:
        y = int(m.group(1)); return f'{y-1}{y}'
    return '기타'

def list_to_sheet(df_raw: pd.DataFrame) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.styles.colors import Color

    df = df_raw.copy()
    rename = {}
    for c in df.columns:
        cl = c.strip().lower()
        if cl in ('품목명','style','item name','product name','description'): rename[c] = '품목명'
        elif cl in ('color','컬러','colour','색상'):  rename[c] = 'Color'
        elif cl in ('사이즈','size','sz'):             rename[c] = '사이즈'
        elif cl in ('현재고','qty','수량','quantity','current qty','currentqty'): rename[c] = '현재고'
        elif cl == 'sku':                               rename[c] = 'SKU'
    df = df.rename(columns=rename)
    required = {'품목명', 'Color', '사이즈'}
    missing = required - set(df.columns)
    if missing: raise ValueError(f"필수 컬럼 없음: {missing}")
    if '현재고' not in df.columns: df['현재고'] = None
    df['현재고'] = pd.to_numeric(df['현재고'], errors='coerce')
    df['품목명'] = df['품목명'].astype(str).str.strip()
    df['Color']  = df['Color'].astype(str).str.strip()
    df['사이즈'] = df['사이즈'].astype(str).str.strip()

    def get_season(sku):
        m = re.search(r'\d{3}[WwSs](\d{2})', str(sku))
        return f'{int(m.group(1))-1}{m.group(1)}' if m else '기타'

    def get_base_code(sku):
        for sz in ['3XL','2XL','XL','FREE','XS','S','M','L']:
            if str(sku).upper().endswith(sz): return str(sku)[:-len(sz)]
        return str(sku)

    if 'SKU' in df.columns:
        df['_시즌'] = df['SKU'].apply(get_season)
        df['_base'] = df['SKU'].apply(get_base_code)
    else:
        df['_시즌'] = '전체'
        df['_base'] = ''

    season_order = sorted([s for s in df['_시즌'].unique() if s != '기타'], reverse=True)
    if '기타' in df['_시즌'].unique(): season_order.append('기타')
    has_qty = df['현재고'].notna().any()

    PALETTE = {
        'V2':    [(0, -0.05), (9,  0.80)],
        'EASY':  [(8,  0.80), (0, -0.05)],
        'ORBAN': [(7,  0.80), (0, -0.05)],
        'SGx28': [(8,  0.80), (0, -0.05)],
        'SGxIG': [(9,  0.80), (0, -0.05)],
        'SGxAP': [(7,  0.80), (0, -0.05)],
        'OTHER': [(5,  0.80), (0, -0.05)],
    }
    SECTION_THEME = {'V2':0,'EASY':8,'ORBAN':7,'SGx28':8,'SGxIG':9,'SGxAP':7,'OTHER':5}
    SECTION_TINT  = 0.40
    HDR_THEME, HDR_TINT = 4, 0.0
    TITLE_THEME, TITLE_TINT = 5, 0.0

    def make_fill(theme_idx, tint):
        c = Color(theme=theme_idx, tint=tint, type='theme')
        return PatternFill(patternType='solid', fgColor=c)

    def get_group(name):
        n = str(name).upper().strip()
        if 'SGX28' in n.replace(' ','') or 'SGX28' in n or 'SGx28' in name: return 'SGx28'
        if 'SGXIG' in n.replace(' ','') or 'SGxIG' in name: return 'SGxIG'
        if 'SGXAP' in n.replace(' ','') or 'SGxAP' in name: return 'SGxAP'
        if n.startswith('V2') or ' V2 ' in n or name.startswith('V2'): return 'V2'
        if 'EASY' in n: return 'EASY'
        if 'ORBAN' in n: return 'ORBAN'
        return 'OTHER'

    def get_category(name):
        n = str(name).upper()
        if any(k in n for k in ['JACKET','ANORAK','BOMBER','PARKA','WIND','VEST','BIKER JAC','DUFFLE']): return 'OUTER'
        if any(k in n for k in ['PANTS','CARGO','BIB','BALLOON','KENDO']): return 'PANTS'
        if any(k in n for k in ['HOODIE','CREWNECK','POLO','SWEAT','ADJUST','LINEWORK','STRAY','WIDE HOOD','ADJUST HOOD']): return 'TOP'
        if any(k in n for k in ['TEE','T-SHIRT','LONG SLEEVE','L/S']): return 'TEE'
        if any(k in n for k in ['BEANIE','CAP','HAT','TRAPPER','BALACLAVA','WARMER']): return 'HAT'
        if any(k in n for k in ['MITTENS','SOCKS','STOMPPAD','STOMP','ACC']): return 'ACC'
        return 'ETC'

    thin = Side(style='thin')
    def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    black_font = Font(name='Arial', size=8, color='000000')
    bold_font  = Font(name='Arial', size=8, color='000000', bold=True)
    center_al  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def write(ws, r, c, val, font=None, fill=None, align=None, border=None):
        cell = ws.cell(r, c, value=val)
        if font:   cell.font   = font
        if fill:   cell.fill   = fill
        if align:  cell.alignment = align
        if border: cell.border = border
        return cell

    wb = Workbook()
    wb.remove(wb.active)
    SIZE_COLS = ['S','M','L','XL','2XL','3XL']

    for season in season_order:
        df_s = df[df['_시즌'] == season].copy()
        if df_s.empty: continue
        all_sz = [s for s in SIZE_COLS if s in df_s['사이즈'].unique()]
        free_sz = df_s['사이즈'].isin({'FREE'}).any()
        ws = wb.create_sheet(title=str(season)[:31])
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 24
        for sz_col in ['E','F','G','H','I','J']: ws.column_dimensions[sz_col].width = 4.5
        ws.column_dimensions['K'].width = 7
        ws.column_dimensions['L'].width = 7
        ws.column_dimensions['M'].width = 13
        ws.merge_cells('A1:M1')
        title_fill = make_fill(TITLE_THEME, TITLE_TINT)
        write(ws, 1, 1, f'SPECIALGUEST® 20{season[-2:] if len(season)==4 else season}FW Season Product-List',
              font=Font(name='Arial', size=10, bold=True, color='000000'),
              fill=title_fill, align=center_al)
        ws.row_dimensions[1].height = 18
        hdr_fill = make_fill(HDR_THEME, HDR_TINT)
        headers = ['CATEGORY','PRODUCT NAME','STYLE NO.','COLOR'] + all_sz + ['C-TOT','S-TOT','BASE PRICE']
        if not all_sz and free_sz:
            headers = ['CATEGORY','PRODUCT NAME','STYLE NO.','COLOR','QTY','','','','','','C-TOT','S-TOT','BASE PRICE']
        for ci, h in enumerate(headers, 1):
            write(ws, 2, ci, h, font=bold_font, fill=hdr_fill, align=center_al, border=bdr())
        ws.row_dimensions[2].height = 14
        cur_row = 3

        def to_en_name(name):
            if name in KR_TO_EN_911: return KR_TO_EN_911[name]
            if re.match(r'^[A-Za-z0-9]', str(name)): return name
            return name
        df_s = df_s.copy()
        df_s['품목명'] = df_s['품목명'].apply(to_en_name)
        style_order = list(dict.fromkeys(df_s['품목명'].tolist()))
        prev_group = None
        grp_item_idx = {}

        for style in style_order:
            grp = get_group(style)
            cat = get_category(style)
            palette = PALETTE[grp]
            if grp not in grp_item_idx: grp_item_idx[grp] = 0
            item_idx = grp_item_idx[grp]
            row_fill = make_fill(*palette[item_idx % 2])
            if grp != prev_group:
                section_names = {
                    'V2':'V2 Collection','EASY':'EASY Collection',
                    'ORBAN':'ORBAN Collection','SGx28':'SPECIALGUEST®  x  RICE28',
                    'SGxIG':'SPECIALGUEST®  x  Insane Garage',
                    'SGxAP':'SPECIALGUEST®  x  Andre Park',
                    'OTHER':'ETC'
                }
                if prev_group is not None: cur_row += 1
                sec_fill = make_fill(SECTION_THEME[grp], SECTION_TINT)
                ws.merge_cells(f'A{cur_row}:M{cur_row}')
                write(ws, cur_row, 1, section_names.get(grp, grp),
                      font=Font(name='Arial', size=9, bold=True, color='000000'),
                      fill=sec_fill, align=left_al)
                ws.row_dimensions[cur_row].height = 14
                cur_row += 1
                prev_group = grp
            grp_item_idx[grp] = item_idx + 1
            sdf = df_s[df_s['품목명'] == style]
            color_order = list(dict.fromkeys(sdf['Color'].tolist()))
            style_first_row = cur_row
            is_free = sdf['사이즈'].isin({'FREE'}).all()

            def base_for_color(color):
                rows = sdf[sdf['Color'] == color]
                if '_base' in rows.columns and rows['_base'].iloc[0]: return rows['_base'].iloc[0]
                return ''

            for ci, color in enumerate(color_order):
                cdf = sdf[sdf['Color'] == color]
                base_code = base_for_color(color)
                pivot = cdf.groupby('사이즈')['현재고'].sum()
                row_total_raw = pivot.sum()
                row_total = int(row_total_raw) if has_qty and pd.notna(row_total_raw) and row_total_raw > 0 else 0
                write(ws, cur_row, 1, cat if ci == 0 else '', font=black_font, fill=row_fill, align=center_al, border=bdr())
                write(ws, cur_row, 2, style if ci == 0 else '', font=black_font, fill=row_fill, align=left_al, border=bdr())
                write(ws, cur_row, 3, base_code, font=black_font, fill=row_fill, align=center_al, border=bdr())
                write(ws, cur_row, 4, color, font=black_font, fill=row_fill, align=left_al, border=bdr())
                if is_free:
                    for col in range(5, 11): write(ws, cur_row, col, '', fill=row_fill, border=bdr())
                    ws.merge_cells(start_row=cur_row, start_column=5, end_row=cur_row, end_column=10)
                    free_raw = pivot.get('FREE', None)
                    free_val = int(free_raw) if has_qty and pd.notna(free_raw) and free_raw > 0 else (0 if has_qty else '')
                    mc = ws.cell(cur_row, 5, value=free_val)
                    mc.font = black_font; mc.fill = row_fill; mc.alignment = center_al; mc.border = bdr()
                else:
                    for si, sz in enumerate(SIZE_COLS, 5):
                        raw = pivot.get(sz, None)
                        val = int(raw) if has_qty and pd.notna(raw) and raw > 0 else (0 if has_qty else '')
                        write(ws, cur_row, si, val, font=black_font, fill=row_fill, align=center_al, border=bdr())
                write(ws, cur_row, 11, f'=SUM(E{cur_row}:J{cur_row})', font=black_font, fill=row_fill, align=center_al, border=bdr())
                write(ws, cur_row, 12, '', font=black_font, fill=row_fill, align=center_al, border=bdr())
                write(ws, cur_row, 13, '', font=black_font, fill=row_fill, align=center_al, border=bdr())
                cur_row += 1

            if len(color_order) > 1:
                ws.merge_cells(f'B{style_first_row}:B{cur_row-1}')
                ws.cell(style_first_row, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.merge_cells(f'A{style_first_row}:A{cur_row-1}')
                ws.cell(style_first_row, 1).alignment = center_al
            s_tot_formula = f'=SUM(K{style_first_row}:K{cur_row-1})'
            if len(color_order) > 1: ws.merge_cells(f'L{style_first_row}:L{cur_row-1}')
            ws.cell(style_first_row, 12).value = s_tot_formula
            ws.cell(style_first_row, 12).font  = black_font
            ws.cell(style_first_row, 12).fill  = row_fill
            ws.cell(style_first_row, 12).alignment = center_al
            ws.cell(style_first_row, 12).border = bdr()

        cur_row += 1
        ws.merge_cells(f'A{cur_row}:D{cur_row}')
        gt_fill = make_fill(HDR_THEME, HDR_TINT)
        write(ws, cur_row, 1, 'GRAND TOTAL', font=bold_font, fill=gt_fill, align=center_al)
        for si in range(5, 12):
            write(ws, cur_row, si, f'=SUM({get_column_letter(si)}3:{get_column_letter(si)}{cur_row-1})', font=bold_font, fill=gt_fill, align=center_al)
        write(ws, cur_row, 12, f'=SUM(L3:L{cur_row-1})', font=bold_font, fill=gt_fill, align=center_al)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def fill_product_sheet(template_file, stock_file):
    from openpyxl.cell.cell import MergedCell as _MC
    stock = pd.read_excel(stock_file)
    stock_col = next((c for c in stock.columns if '911' in str(c) or '현재고' in str(c)), stock.columns[-1])

    def get_base_sz(sku):
        for sz in ['3XL','2XL','XL','FREE','XS','S','M','L']:
            if str(sku).upper().endswith(sz): return str(sku)[:-len(sz)].upper(), sz.upper()
        return str(sku).upper(), 'FREE'

    lookup = {}
    for _, row in stock.iterrows():
        base, sz = get_base_sz(str(row['자사코드']))
        qty = row[stock_col]
        lookup[(base, sz)] = int(qty) if pd.notna(qty) else 0

    tpl_bytes = template_file.read() if hasattr(template_file, 'read') else open(template_file,'rb').read()
    wb = load_workbook(io.BytesIO(tpl_bytes))
    total_filled = 0
    for ws in wb.worksheets:
        header_row = None
        size_col_map = {}
        for r in range(1, min(ws.max_row+1, 10)):
            for c in range(1, ws.max_column+1):
                v = ws.cell(r, c).value
                if v and str(v).strip().upper() == 'STYLE NO.':
                    header_row = r; break
            if header_row: break
        if not header_row: continue
        for c in range(1, ws.max_column+1):
            v = ws.cell(header_row, c).value
            if v and str(v).strip().upper() in ['S','M','L','XL','2XL','3XL']:
                size_col_map[str(v).strip().upper()] = c
        for r in range(header_row+1, ws.max_row+1):
            c3 = ws.cell(r, 3)
            if isinstance(c3, _MC) or not c3.value: continue
            base = str(c3.value).strip().upper()
            for sz, col in size_col_map.items():
                target = ws.cell(r, col)
                if not isinstance(target, _MC):
                    qty = lookup.get((base, sz), 0)
                    target.value = qty if qty > 0 else 0
                    if qty > 0: total_filled += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total_filled

def sheet_to_list(xlsx_file, base_csv: pd.DataFrame = None) -> pd.DataFrame:
    wb = load_workbook(xlsx_file, data_only=True); ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None; header = []
    for i, row in enumerate(rows):
        row_upper = [str(c).strip().upper() if c is not None else '' for c in row]
        if any(k in row_upper for k in ('품목명','STYLE','PRODUCT NAME','COLOR')):
            header_idx = i; header = row_upper; break
    if header_idx is None: raise ValueError("헤더 행을 찾을 수 없습니다.")
    style_col = next((i for i,h in enumerate(header) if h in ('품목명','STYLE','PRODUCT NAME','DESCRIPTION')), None)
    color_col  = next((i for i,h in enumerate(header) if h in ('COLOR','COLOUR','컬러','색상')), None)
    if style_col is None or color_col is None: raise ValueError(f"품목명/Color 컬럼 없음. 헤더: {header}")
    skip_keywords = {'TOTAL','C-TOT','S-TOT','GRAND','BASE PRICE','PRICE','RETAIL','SUPPLY','CHINA','CATEGORY','PRODUCT','STYLE','COLOR','COLOUR','품목명','컬러',''}
    size_col_map = {}
    for i, h in enumerate(header):
        if h in skip_keywords: continue
        if h in [s.upper() for s in SIZE_ORDER_LIST]: size_col_map[i] = h
        elif len(h) <= 5 and i > color_col and 'TOT' not in h and 'PRICE' not in h: size_col_map[i] = h
    if not size_col_map: raise ValueError(f"사이즈 컬럼 없음. 헤더: {header}")
    records = []; last_style = ''
    for row in rows[header_idx+1:]:
        if all(c is None or str(c).strip() == '' for c in row): continue
        style_val = row[style_col] if style_col < len(row) else None
        color_val  = row[color_col]  if color_col  < len(row) else None
        if style_val and str(style_val).strip():
            s = str(style_val).strip()
            if s.upper() not in ('TOTAL','GRAND TOTAL',''): last_style = s
        style_val = last_style
        if not style_val or not color_val or str(color_val).strip() == '': continue
        if str(color_val).strip().upper() in ('TOTAL','GRAND TOTAL','COLOR'): continue
        color_val = str(color_val).strip()
        for col_idx, size_name in size_col_map.items():
            if col_idx >= len(row): continue
            cell_val = row[col_idx]
            if cell_val is None or str(cell_val).strip() == '': qty = None
            else:
                try: qty = int(float(str(cell_val)))
                except (ValueError, TypeError): continue
            records.append({'품목명':style_val,'Color':color_val,'사이즈':size_name,'현재고':qty})
    if not records: raise ValueError("변환된 데이터가 없습니다.")
    df_sheet = pd.DataFrame(records)
    if base_csv is not None:
        df_base = base_csv.copy()
        col_rename = {}
        for c in df_base.columns:
            cl = c.strip().lower()
            if cl in ('품목명','style','item name'): col_rename[c] = '품목명'
            elif cl in ('color','컬러','colour'): col_rename[c] = 'Color'
            elif cl in ('사이즈','size','sz'): col_rename[c] = '사이즈'
            elif cl in ('현재고','qty','수량','current qty'): col_rename[c] = '현재고'
            elif cl == 'sku': col_rename[c] = 'SKU'
        df_base = df_base.rename(columns=col_rename)
        sheet_lookup = {(str(r['품목명']).strip(), str(r['Color']).strip(), str(r['사이즈']).strip()): r['현재고'] for _, r in df_sheet.iterrows()}
        df_base['현재고'] = df_base.apply(lambda row: sheet_lookup.get((str(row['품목명']).strip(), str(row['Color']).strip(), str(row['사이즈']).strip()), None), axis=1)
        out_cols = [c for c in ['SKU','품목명','Color','사이즈','현재고'] if c in df_base.columns]
        return df_base[out_cols]
    return df_sheet[['품목명','Color','사이즈','현재고']]


# ══════════════════════════════════════════════════════
# TAB 3: 911스포츠 입고 출고서
# ══════════════════════════════════════════════════════

KR_TO_EN_911 = {
    '2526 스페셜게스트 브이투 에어박스 미드레이어 다운 보드복 자켓':   '108. V2 AIRBOX MIDLAYER DOWN JACKET',
    '2526 스페셜게스트 브이투 에어패널 미드레이어 다운 아노락 보드복 자켓': '109. V2 AIRPANEL MIDLAYER DOWN ANORAK',
    '2526 스페셜게스트 브이투 페더플로우 윈드 보드복 자켓':          '110. V2 FEATHERFLOW WIND JACKET',
    '2526 스페셜게스트 브이투 모드 씬슬레이트 아노락 보드복 자켓':    '70. V2 MOD THINSULATE ANORAK JACKET',
    '2526 스페셜게스트 브이투 모드 슈퍼와이드 카고 보드복 팬츠':     '71. V2 MOD SUPERWIDE CARGO PANTS',
    '2526 스페셜게스트 브이투 사이드라인 트랙 보드복 자켓':          '40. V2 SIDELINE TRACK JACKET',
    '2526 스페셜게스트 브이투 사이드라인 슈퍼와이드 보드복 팬츠':    '41. V2 SIDELINE SUPERWIDE PANTS',
    '2526 스페셜게스트 브이투 와이드 바시티 보드복 자켓':            '42. V2 WIDE VARSITY JACKET',
    '2526 스페셜게스트 브이투 아노락 FC 보드복 자켓':               '01. V2 ANORAK JACKET FC',
    '2526 스페셜게스트 브이투 오버와이드 카고 FC 보드복 팬츠':       '05. V2 OVERWIDE CARGO PANTS FC',
    '2526 스페셜게스트 브이투 윈드 브레이커 R 보드복 자켓':          '04. V2 WIND BREAKER R',
    '2526 스페셜게스트 브이투 어저스트 후드':                        '08. V2 ADJUST HOODIE N',
    '2526 스페셜게스트 브이투 롱 슬리브':                            '09. V2 LONG SLEEVE R',
    '2526 스페셜게스트 브이투 핀스트라이프 크루넥':                  '91. V2 PINSTRIPE CREWNECK',
    '2526 스페셜게스트 브이투 알터에고 머서라이즈드 티셔츠':         '95. V2 ALTEREGO MERCERIZED TEE',
    '2526 스페셜게스트 브이투 알터에고 머서라이즈드 긴팔 티셔츠':    '96. V2 ALTEREGO MERCERIZED L/S TEE',
    '2526 스페셜게스트 브이투 라인플로우 머서라이즈드 긴팔 티셔츠':  '97. V2 LINEFLOW MERCERIZED L/S TEE',
    '2526 스페셜게스트 브이투 캔디 비니':                            '61. V2 CANDY BEANIE',
    '2526 스페셜게스트 브이투 스트로크 자카드 비니':                 '75. V2 STROKE JACQUARD BEANIE',
    '2526 스페셜게스트 브이투 유니버셜 볼캡':                        '62. V2 UNIVERSAL BALL CAP',
    '2526 스페셜게스트 브이투 클라우드 퍼 비니':                     '83. V2 CLOUD FUR BEANIE',
    '2526 스페셜게스트 이지 버스트라인 봄버 보드복 자켓':            '105. EASY BURSTLINE BOMBER JACKET',
    '2526 스페셜게스트 이지 울트라와이드 켄도 보드복 팬츠':          '106. EASY ULTRAWIDE KENDO PANTS',
    '2526 스페셜게스트 이지 마크스티치 워크 보드복 자켓':            '107. EASY MARKSTITCH WORK JACKET',
    '2526 스페셜게스트 이지 보드복 팬츠':                            '16. EASY PANTS N',
    '2526 스페셜게스트 이지 슈퍼와이드 블럭 보드복 팬츠':            '47. EASY SUPERWIDE BLOCK PANTS',
    '2526 스페셜게스트 이지 라인워크 크루넥':                        '18. EASY LINEWORK CREWNECK',
    '2526 스페셜게스트 이지 스트레이 스트라이프 집 폴로 셔츠':       '19. EASY STRAY STRIPED ZIP POLO',
    '2526 스페셜게스트 이지 와이드 후드':                            '55. EASY WIDE HOODIE',
    '2526 스페셜게스트 이지 셰르파 트래퍼 캡':                       '22. EASY SHERPA TRAPPER CAP',
    '2526 스페셜게스트 이지 비니':                                   '63. EASY BEANIE R',
    '2526 스페셜게스트 이지 뉴 웨이브 비니':                         '64. EASY NEW WAVE BEANIE',
    '2526 스페셜게스트 이지 와이드 볼캡':                            '65. EASY WIDE BALL CAP',
    '2526 스페셜게스트 오반 엑티브 보드복 자켓':                     '53. ORBAN ACTIVE JACKET R',
    '2526 스페셜게스트 오반 엑티브 보드복 팬츠':                     '29. ORBAN ACTIVE PANTS R',
    '2526 스페셜게스트 오반 아웃쉘 보드복 자켓':                     '48. ORBAN OUTSHELL JACKET 3Layer',
    '2526 스페셜게스트 오반 오버와이드 빕 보드복 팬츠':              '49. ORBAN OVERWIDE BIB 3Layer',
    '2526 스페셜게스트 오반 자카드 비니':                            '77. ORBAN JACQUARD BEANIE',
    '2526 스페셜게스트X라이스28 이지 바이커 보드복 자켓':            '101. SGx28 EASY BIKER JACKET',
    '2526 스페셜게스트X라이스28 이지 울트라와이드 바이커 보드복 팬츠':'102. SGx28 EASY ULTRAWIDE BIKER PANTS',
    '2526 스페셜게스트X라이스28 브이투 어저스트 크루넥':             '90. SGx28 V2 ADJUST CREWNECK',
    '2526 스페셜게스트X라이스28 이지 싱크 자카드 비니':              '81. SGx28 EASY SYNC JACQUARD BEANIE',
    '2526 스페셜게스트X인세인개러지 인세인 코드 보드복 자켓':         '111. SGxIG INSANE CODE JACKET',
    '2526 스페셜게스트X인세인개러지 슈퍼와이드 인세인 코드 보드복 팬츠':'112. SGxIG SUPERWIDE INSANE CODE PANTS',
    '2526 스페셜게스트X인세인개러지 페트 개러지 보드복 자켓':         '113. SGxIG FAITE GARAGE JACKET',
    '2526 스페셜게스트X인세인개러지 슈퍼와이드 페트 개러지 보드복 팬츠':'114. SGxIG SUPERWIDE FAITE GARAGE PANTS',
    '2526 스페셜게스트X앙드레박 이지 더플 보드복 자켓':              '103. SGxAP EASY DUFFLE JACKET',
    '2526 스페셜게스트X앙드레박 이지 오버와이드 벌룬 보드복 팬츠':   '104. SGxAP EASY OVERWIDE BALLOON PANTS',
    '2526 스페셜게스트X앙드레박 이지 루프라인 트래퍼 햇':            '92. SGxAP EASY LOOPLINE TRAPPER HAT',
}

SIZE_COLS_911 = ['S','M','L','XL','2XL','3XL']
GROUP_MAIN_COLORS_911 = {'V2':'E2EFDA','EASY':'DDEBF7','ORBAN':'FFF2CC','COLLAB':'FCE4D6','OTHER':'F2F2F2'}
GROUP_STOT_COLORS_911 = {'V2':'C6E0B4','EASY':'BDD7EE','ORBAN':'FFD966','COLLAB':'F4B183','OTHER':'D9D9D9'}

def _get_group_911(n):
    u = str(n).upper()
    if 'SGX' in u: return 'COLLAB'
    if 'V2 ' in u or 'V2\t' in u: return 'V2'
    if 'EASY' in u: return 'EASY'
    if 'ORBAN' in u: return 'ORBAN'
    return 'OTHER'

def _extr_color(name):
    parts = str(name).strip().split(); cp = []
    for p in reversed(parts):
        if re.match(r'^[A-Za-z]', p): cp.insert(0, p)
        else: break
    return ' '.join(cp)

def _extr_base(name):
    c = _extr_color(name); b = str(name).strip()
    return b[:b.rfind(c)].strip() if c else b

SIZE_TO_COL_911 = {'S':3, 'M':4, 'L':5, 'XL':6, '2XL':7, '3XL':8}
HW_MERGE_START = 4
HW_MERGE_END   = 7

def _normalize_color(s):
    return re.sub(r'[\s\-/]', '', str(s).lower().strip())

def _get_base(sku):
    for sz in ['3XL', '2XL', 'XL', 'FREE', 'XS', 'S', 'M', 'L']:
        if str(sku).upper().endswith(sz):
            return str(sku)[:-len(sz)].upper()
    return str(sku).upper()

def make_restock_output(stock_file, template_file, hq_csv_file=None, target_qty=2, hq_min=1):
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell as _MergedCell
    from openpyxl.utils import get_column_letter

    stock = pd.read_excel(stock_file)
    stock['컬러']       = stock['상품명'].apply(_extr_color)
    stock['제품그룹']   = stock['상품명'].apply(_extr_base)
    stock['사이즈']     = stock['옵션명'].fillna('FREE').astype(str).str.strip()
    stock['현재고']     = stock['[매장] 오프라인_911스포츠'].fillna(0).astype(int)
    stock['영문품목명'] = stock['제품그룹'].map(KR_TO_EN_911)

    hq_lookup = {}
    if hq_csv_file is not None:
        df_hq = pd.read_csv(hq_csv_file, encoding='utf-8-sig')
        df_hq['현재고'] = pd.to_numeric(df_hq['현재고'], errors='coerce').fillna(0).astype(int)
        for _, r in df_hq.iterrows():
            hq_lookup[(_get_base(r['SKU']), str(r['Size']).strip().upper())] = r['현재고']

    def calc_shipout(row):
        q911 = row['현재고']
        sz   = str(row['사이즈']).strip().upper()
        base = _get_base(str(row['자사코드']))
        if q911 >= target_qty:
            return 0
        needed = target_qty - q911
        if hq_lookup:
            available = max(0, hq_lookup.get((base, sz), 0) - hq_min)
            return min(needed, available)
        else:
            return needed

    stock['출고수량'] = stock.apply(calc_shipout, axis=1)

    qty_lookup   = {}
    price_lookup = {}
    for _, row in stock.iterrows():
        if pd.isna(row['영문품목명']): continue
        en = str(row['영문품목명']).strip()
        nc = _normalize_color(row['컬러'])
        sz = str(row['사이즈']).strip()
        qty_lookup[(en, nc, sz)] = int(row['출고수량'])
        if (en, nc) not in price_lookup and pd.notna(row.get('소비자가', None)):
            price_lookup[(en, nc)] = int(row['소비자가'])

    if template_file is None:
        return io.BytesIO(), stock

    tpl_bytes = template_file.read() if hasattr(template_file, 'read') else open(template_file, 'rb').read()
    wb_tpl = load_workbook(io.BytesIO(tpl_bytes))
    ws_tpl = wb_tpl.active

    b_rows = []
    for row in ws_tpl.iter_rows(min_row=3, max_row=ws_tpl.max_row, values_only=False):
        b = row[1]
        if b.value and str(b.value).strip():
            b_rows.append(b.row)

    tpl_blocks = []
    for i, start_row in enumerate(b_rows):
        end_row = b_rows[i+1] - 1 if i+1 < len(b_rows) else ws_tpl.max_row
        prod_name = str(ws_tpl.cell(start_row, 2).value).strip()
        colors_in_tpl = []
        for r in range(start_row, end_row+1):
            d = ws_tpl.cell(r, 4).value
            if d and str(d).strip():
                colors_in_tpl.append(str(d).strip())
        tpl_blocks.append({
            'prod': prod_name,
            'ref_row': start_row,
            'colors_tpl': colors_in_tpl,
        })

    wb_new = load_workbook(io.BytesIO(tpl_bytes))
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    ws_src = wb_new.active

    for col, cd in ws_src.column_dimensions.items():
        if cd.width: ws_out.column_dimensions[col].width = cd.width

    from copy import copy

    def copy_cell_style(src_cell, dst_cell):
        if isinstance(src_cell, _MergedCell): return
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format

    for r in [1, 2]:
        ws_out.row_dimensions[r].height = ws_src.row_dimensions[r].height or 15
        for c in range(1, 16):
            src_cell = ws_src.cell(r, c)
            dst_cell = ws_out.cell(r, c)
            if not isinstance(src_cell, _MergedCell):
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)
    for merge in ws_src.merged_cells.ranges:
        if merge.min_row <= 2:
            ws_out.merge_cells(
                start_row=merge.min_row, start_column=merge.min_col,
                end_row=merge.max_row,   end_column=merge.max_col
            )

    cur_row = 3

    for block in tpl_blocks:
        prod    = block['prod']
        ref_row = block['ref_row']

        colors_out = []
        for color in block['colors_tpl']:
            nc = _normalize_color(color)
            is_free = qty_lookup.get((prod, nc, 'FREE'), -1) >= 0
            if is_free:
                if qty_lookup.get((prod, nc, 'FREE'), 0) > 0:
                    colors_out.append(color)
            else:
                total = sum(qty_lookup.get((prod, nc, sz), 0) for sz in SIZE_TO_COL_911)
                if total > 0:
                    colors_out.append(color)

        if not colors_out:
            continue

        prod_first_row = cur_row
        n_colors = len(colors_out)

        for ci, color in enumerate(colors_out):
            nc    = _normalize_color(color)
            price = price_lookup.get((prod, nc))
            src_r = ref_row if ci == 0 else min(ref_row + 1, ws_src.max_row)

            ws_out.row_dimensions[cur_row].height = ws_src.row_dimensions[ref_row].height or 15
            for col in range(1, 16):
                src_cell = ws_src.cell(src_r, col)
                dst_cell = ws_out.cell(cur_row, col)
                copy_cell_style(src_cell, dst_cell)

            ws_out.cell(cur_row, 2).value = prod if ci == 0 else None
            ws_out.cell(cur_row, 3).value = ws_src.cell(ref_row, 3).value if ci == 0 else None
            ws_out.cell(cur_row, 4).value = color

            is_free_prod = qty_lookup.get((prod, nc, 'FREE'), -1) >= 0

            if is_free_prod:
                qty = qty_lookup.get((prod, nc, 'FREE'), 0)
                for col in range(5, 11): ws_out.cell(cur_row, col).value = None
                ws_out.merge_cells(start_row=cur_row, start_column=HW_MERGE_START,
                                   end_row=cur_row,   end_column=HW_MERGE_END)
                mc = ws_out.cell(cur_row, HW_MERGE_START)
                mc.value = qty
                copy_cell_style(ws_src.cell(src_r, HW_MERGE_START), mc)
                mc.alignment = Alignment(horizontal='center', vertical='center')
            else:
                size_col_map_out = {'S':5,'M':6,'L':7,'XL':8,'2XL':9,'3XL':10}
                for sz, col in size_col_map_out.items():
                    qty = qty_lookup.get((prod, nc, sz), 0)
                    ws_out.cell(cur_row, col).value = qty if qty > 0 else 0

            ws_out.cell(cur_row, 11).value = f'=SUM(E{cur_row}:J{cur_row})'
            copy_cell_style(ws_src.cell(src_r, 11), ws_out.cell(cur_row, 11))

            ws_out.cell(cur_row, 12).value = None
            copy_cell_style(ws_src.cell(src_r, 12), ws_out.cell(cur_row, 12))

            if price and ci == 0:
                ws_out.cell(cur_row, 13).value = price
                copy_cell_style(ws_src.cell(ref_row, 13), ws_out.cell(cur_row, 13))

            if ci == 0:
                ws_out.cell(cur_row, 14).value = f'=SUM(M{cur_row}*0.7)'
                copy_cell_style(ws_src.cell(ref_row, 14), ws_out.cell(cur_row, 14))

            if ci == 0:
                ws_out.cell(cur_row, 15).value = f'=SUM(L{cur_row}*N{cur_row})'
                copy_cell_style(ws_src.cell(ref_row, 15), ws_out.cell(cur_row, 15))

            cur_row += 1

        if n_colors > 1:
            ws_out.merge_cells(start_row=prod_first_row, start_column=2,
                               end_row=cur_row-1,        end_column=2)
            ws_out.cell(prod_first_row, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws_out.merge_cells(start_row=prod_first_row, start_column=3,
                               end_row=cur_row-1,        end_column=3)

        j_formula = f'=SUM(K{prod_first_row}:K{cur_row-1})'
        if n_colors > 1:
            ws_out.merge_cells(start_row=prod_first_row, start_column=12,
                               end_row=cur_row-1,        end_column=12)
        ws_out.cell(prod_first_row, 12).value = j_formula
        copy_cell_style(ws_src.cell(ref_row, 12), ws_out.cell(prod_first_row, 12))
        ws_out.cell(prod_first_row, 12).alignment = Alignment(horizontal='center', vertical='center')

        if n_colors > 1:
            first_nc = _normalize_color(colors_out[0])
            all_same_price = all(
                price_lookup.get((prod, _normalize_color(c))) == price_lookup.get((prod, first_nc))
                for c in colors_out
            )
            if all_same_price:
                for kcol in [13, 14, 15]:
                    ws_out.merge_cells(start_row=prod_first_row, start_column=kcol,
                                       end_row=cur_row-1,        end_column=kcol)
                    ws_out.cell(prod_first_row, kcol).alignment = Alignment(horizontal='center', vertical='center')

    total_row = cur_row + 1
    ws_out.cell(total_row-1, 11).value = 'Quantity'
    ws_out.cell(total_row-1, 13).value = 'Total supply price(KRW)'
    ws_out.cell(total_row, 11).value   = f'=SUM(K3:K{cur_row-1})'
    ws_out.cell(total_row, 12).value   = f'=SUM(L3:L{cur_row-1})'
    ws_out.cell(total_row, 15).value   = f'=SUM(O3:O{cur_row-1})'

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf, stock


# ════════════════════════════════════════════════════════
# TAB 4: 💱 국가별 가격 생성기
# ════════════════════════════════════════════════════════

def make_country_pricing_sheet(product_file, country_list, exchange_rates):
    """프로덕트 시트에서 기본가격을 읽고 국가별 가격 시트 생성 (통합본)"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    prod_bytes = product_file.read() if hasattr(product_file, 'read') else open(product_file, 'rb').read()
    wb_prod = load_workbook(io.BytesIO(prod_bytes), data_only=True)
    ws_prod = wb_prod.active
    
    products = []
    for row in range(4, ws_prod.max_row + 1):
        cat = ws_prod.cell(row, 1).value
        prod_name = ws_prod.cell(row, 2).value
        style_no = ws_prod.cell(row, 3).value
        color = ws_prod.cell(row, 4).value
        price_krw = ws_prod.cell(row, 13).value
        
        if prod_name and style_no and price_krw:
            try:
                price_krw = float(price_krw)
                products.append({
                    'category': cat,
                    'product_name': prod_name,
                    'style_no': style_no,
                    'color': color,
                    'price_krw': price_krw,
                })
            except:
                pass
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Country Pricing'
    
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = ['Category', 'Product Name', 'Style No.', 'Color', 'Base Price (KRW)']
    headers += [f"{c['code']} ({c['symbol']})" for c in country_list]
    
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006FC0')
        cell.alignment = center_align
        cell.border = border_all
    
    ws_out.column_dimensions['A'].width = 15
    ws_out.column_dimensions['B'].width = 40
    ws_out.column_dimensions['C'].width = 16
    ws_out.column_dimensions['D'].width = 20
    ws_out.column_dimensions['E'].width = 18
    
    for col in range(6, 6 + len(country_list)):
        ws_out.column_dimensions[get_column_letter(col)].width = 16
    
    for row_idx, prod in enumerate(products, 2):
        cell = ws_out.cell(row=row_idx, column=1, value=prod['category'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=2, value=prod['product_name'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=3, value=prod['style_no'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=4, value=prod['color'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=5, value=int(prod['price_krw']))
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0'
        
        for col_idx, country in enumerate(country_list, 6):
            code = country['code']
            adjustment = country['adjustment']
            exchange = exchange_rates.get(code, 1.0)
            
            price_local = (prod['price_krw'] / exchange) * (1 + adjustment)
            
            cell = ws_out.cell(row=row_idx, column=col_idx, value=price_local)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = center_align
            cell.border = border_all
            cell.number_format = '#,##0.00'
    
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


def make_single_country_sheet(product_file, country, exchange_rate):
    """단일 국가별 가격 시트 생성 (개별 다운로드용)"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    prod_bytes = product_file.read() if hasattr(product_file, 'read') else open(product_file, 'rb').read()
    wb_prod = load_workbook(io.BytesIO(prod_bytes), data_only=True)
    ws_prod = wb_prod.active
    
    products = []
    for row in range(4, ws_prod.max_row + 1):
        cat = ws_prod.cell(row, 1).value
        prod_name = ws_prod.cell(row, 2).value
        style_no = ws_prod.cell(row, 3).value
        color = ws_prod.cell(row, 4).value
        price_krw = ws_prod.cell(row, 13).value
        
        if prod_name and style_no and price_krw:
            try:
                price_krw = float(price_krw)
                products.append({
                    'category': cat,
                    'product_name': prod_name,
                    'style_no': style_no,
                    'color': color,
                    'price_krw': price_krw,
                })
            except:
                pass
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = country['code']
    
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = ['Category', 'Product Name', 'Style No.', 'Color', 'Base Price (KRW)', f"{country['code']} ({country['symbol']})"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006FC0')
        cell.alignment = center_align
        cell.border = border_all
    
    ws_out.column_dimensions['A'].width = 15
    ws_out.column_dimensions['B'].width = 40
    ws_out.column_dimensions['C'].width = 16
    ws_out.column_dimensions['D'].width = 20
    ws_out.column_dimensions['E'].width = 18
    ws_out.column_dimensions['F'].width = 18
    
    for row_idx, prod in enumerate(products, 2):
        cell = ws_out.cell(row=row_idx, column=1, value=prod['category'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=2, value=prod['product_name'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=3, value=prod['style_no'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=4, value=prod['color'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=5, value=int(prod['price_krw']))
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0'
        
        adjustment = country['adjustment']
        price_local = (prod['price_krw'] / exchange_rate) * (1 + adjustment)
        
        cell = ws_out.cell(row=row_idx, column=6, value=price_local)
        cell.font = Font(name='Arial', size=9, bold=True, color='006FC0')
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0.00'
    
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════
# MAIN UI
# ════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4 = st.tabs(["📦 패킹 문서 생성기", "🔄 시트 변환기", "🏪 911스포츠 입고 출고서", "💱 국가별 가격 생성기"])

# ── TAB 1
with tab1:
    st.caption("해외 출고리스트 업로드 → 패킹리스트 + 인보이스 + Actual Packing List + 카테고리별 시트 자동 생성")
    st.divider()
    col1, col2 = st.columns([1,1])
    with col1:
        st.subheader("📁 파일 업로드")
        csv_file = st.file_uploader("해외 출고리스트 (.csv)", type=['csv'], key="tab1_upload")
    with col2:
        st.subheader("⚙️ 설정")
        messrs      = st.text_input("거래처명 (MESSRS)", placeholder="예: 傲刃堂（成都）...", key="messrs")
        destination = st.text_input("Final Destination", value="China", key="dest")
        date_str    = st.text_input("DATE", value=datetime.now().strftime("%Y/%m/%d"), key="date")
    st.divider()
    if csv_file:
        with st.spinner("처리 중..."):
            try:
                df, boxes = parse_export_csv(csv_file)
                total_qty    = df['수량'].sum(); total_boxes  = len(boxes)
                total_weight = df.groupby('박스번호')['무게(kg)'].first().sum()
                total_cbm    = round(total_boxes * CBM_PER_BOX, 3)
                st.success(f"✅ 파싱 완료 — **{total_boxes} CTNs / {total_qty} pcs / {total_weight}kg / CBM {total_cbm}**")
                with st.expander("📊 카테고리별 요약"):
                    cat_summary = df.groupby('Category')['수량'].sum().sort_index()
                    for cat, qty in cat_summary.items(): st.write(f"**{cat}**: {qty}pcs")
                st.divider(); st.subheader("📥 다운로드")
                c1, c2, c3, c4 = st.columns(4)
                with c1: st.download_button("⬇️ 패킹리스트", data=make_packing_list(boxes, destination), file_name=f"PackingList_{date_str.replace('/','')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                with c2: st.download_button("⬇️ 인보이스", data=make_invoice(df, messrs, destination, date_str), file_name=f"Invoice_{date_str.replace('/','')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                with c3: st.download_button("⬇️ Actual Packing List", data=make_actual_packing_list(boxes, messrs, destination, date_str), file_name=f"ActualPackingList_{date_str.replace('/','')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                with c4: st.download_button("⬇️ 카테고리별", data=make_category_packing_list(boxes), file_name=f"CategoryPackingList_{date_str.replace('/','')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e:
                st.error(f"❌ 오류: {e}")
                import traceback; st.code(traceback.format_exc())
    else:
        st.info("👆 해외 출고리스트 CSV 파일을 업로드해주세요.")

# ── TAB 2
with tab2:
    st.caption("제품 리스트 ↔ 제품 시트 양방향 변환")
    sub_tab1, sub_tab2 = st.tabs(["📋 리스트 → 시트", "📊 시트 → 리스트"])
    with sub_tab1:
        st.caption("제품 시트 템플릿 + 재고 파일 → STYLE NO. 기준으로 수량 자동 입력")
        st.divider()
        col_a, col_b = st.columns(2)
        with col_a:
            l2s_tpl = st.file_uploader("① 제품 시트 템플릿 (.xlsx)", type=['xlsx'], key="l2s_tpl")
            st.caption("📌 C열에 STYLE NO.(베이스코드) 있는 시트")
        with col_b:
            l2s_stock = st.file_uploader("② 재고 파일 (.xlsx)", type=['xlsx'], key="l2s_stock")
            st.caption("📌 자사코드 + [매장] 오프라인_911스포츠 컬럼 있는 파일")
        if l2s_tpl and l2s_stock:
            with st.spinner("수량 채우는 중..."):
                try:
                    result_buf, filled_count = fill_product_sheet(l2s_tpl, l2s_stock)
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    tpl_name = l2s_tpl.name.replace('.xlsx','')
                    st.success(f"✅ 완료 — {filled_count}개 셀에 수량 입력됨")
                    st.download_button("⬇️ 수량 채워진 제품 시트 다운로드", data=result_buf,
                        file_name=f"{tpl_name}_재고입력_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary")
                except Exception as e:
                    st.error(f"❌ 오류: {e}"); import traceback; st.code(traceback.format_exc())
        elif l2s_tpl and not l2s_stock: st.warning("② 재고 파일도 업로드해주세요.")
        elif l2s_stock and not l2s_tpl: st.warning("① 제품 시트 템플릿도 업로드해주세요.")
        else: st.info("👆 제품 시트 템플릿과 재고 파일을 모두 업로드해주세요.")
    with sub_tab2:
        st.markdown("**제품 시트 Excel** → **제품리스트 CSV**\n시트에서 현재고를 읽어, 기존 제품리스트의 `현재고` 컬럼을 채워서 반환합니다.")
        st.divider()
        col_a, col_b = st.columns(2)
        with col_a: s2l_sheet = st.file_uploader("① 제품 시트 Excel (.xlsx)  ← 현재고가 채워진 시트", type=['xlsx'], key="s2l_sheet")
        with col_b: s2l_base  = st.file_uploader("② 기존 제품리스트 CSV (선택) ← SKU 유지용", type=['csv','xlsx'], key="s2l_base")
        if s2l_sheet:
            try:
                df_base = None
                if s2l_base:
                    if s2l_base.name.endswith('.csv'): df_base = pd.read_csv(s2l_base, encoding='utf-8-sig')
                    else: df_base = pd.read_excel(s2l_base)
                    st.info(f"📎 기존 리스트 로드됨 — {len(df_base):,}행 (SKU 매핑 적용)")
                if st.button("▶️ 리스트로 변환", key="s2l_btn", use_container_width=True, type="primary"):
                    with st.spinner("변환 중..."):
                        df_result = sheet_to_list(s2l_sheet, base_csv=df_base)
                        total_filled = df_result['현재고'].notna().sum(); total_rows = len(df_result)
                        total_qty_s = int(df_result['현재고'].fillna(0).sum())
                        st.success(f"✅ 변환 완료 — {total_rows:,}행 / 현재고 매핑 {total_filled:,}건 / 총 {total_qty_s:,}pcs")
                        with st.expander("미리보기 (상위 20행)"): st.dataframe(df_result.head(20), use_container_width=True)
                        filled = df_result[df_result['현재고'].notna() & (df_result['현재고'] > 0)]
                        if not filled.empty:
                            with st.expander("📊 품목별 수량 요약"):
                                summary = filled.groupby('품목명')['현재고'].sum().reset_index(); summary.columns = ['품목명','수량합계']
                                st.dataframe(summary, use_container_width=True)
                        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                        csv_out = df_result.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button("⬇️ 제품리스트 CSV 다운로드", data=csv_out.encode('utf-8-sig'),
                            file_name=f"제품리스트_{ts}.csv", mime="text/csv", use_container_width=True)
            except Exception as e:
                st.error(f"❌ 오류: {e}"); import traceback; st.code(traceback.format_exc())
        else: st.info("👆 현재고가 채워진 제품 시트 Excel을 업로드해주세요.")

# ── TAB 3: 911스포츠 입고 출고서
with tab3:
    st.caption("911 매장재고 + 본사재고 기준으로 부족분 자동 계산 → 입고 양식 출력")
    st.divider()

    col1, col2 = st.columns([2, 1])
    with col1:
        st.subheader("📁 파일 업로드")
        template_file_911 = st.file_uploader(
            "① 입고 양식 Excel (.xlsx) — 원본 그대로 업로드",
            type=['xlsx'], key="tab3_template"
        )
        stock_file_911 = st.file_uploader(
            "② 매장 재고현황 Excel (.xlsx)",
            type=['xlsx'], key="tab3_stock"
        )
        st.caption("📌 컬럼 필수: 자사코드 / 상품명 / 옵션명 / **[매장] 오프라인_911스포츠**")
        hq_file_911 = st.file_uploader(
            "③ 본사 재고 CSV (.csv)",
            type=['csv'], key="tab3_hq"
        )
        st.caption("📌 컬럼 필수: SKU / Size / 현재고  (재고 없는 항목은 빈 값도 OK)")
    with col2:
        st.subheader("⚙️ 설정")
        target_qty_911 = st.number_input(
            "목표 재고 수량", min_value=1, max_value=10, value=2, step=1,
            key="tab3_target",
            help="각 SKU를 이 수량까지 채우는 것을 목표로 합니다"
        )
        hq_min_911 = st.number_input(
            "본사 최소 잔류 수량", min_value=0, max_value=10, value=1, step=1,
            key="tab3_hq_min",
            help="출고 후 본사에 최소 이 수량은 남겨둡니다"
        )

    st.divider()

    if stock_file_911 and template_file_911:
        with st.spinner("재고 분석 및 양식 채우는 중..."):
            try:
                buf_911, df_911 = make_restock_output(
                    stock_file_911, template_file_911,
                    hq_csv_file=hq_file_911,
                    target_qty=target_qty_911,
                    hq_min=hq_min_911
                )
                df_need        = df_911[df_911['출고수량'] > 0]
                df_full        = df_911[df_911['출고수량'] == 0]
                total_out_911  = int(df_need['출고수량'].sum())
                sku_out_count  = len(df_need)
                sku_full_count = len(df_full)

                c1, c2, c3 = st.columns(3)
                c1.metric("총 출고 수량", f"{total_out_911}pcs")
                c2.metric("출고 필요 SKU", f"{sku_out_count}건")
                c3.metric("재고 충족 SKU", f"{sku_full_count}건")

                with st.expander("📊 품목별 출고 수량 요약"):
                    summary_911 = (
                        df_need.groupby('영문품목명')['출고수량']
                        .sum().reset_index()
                        .rename(columns={'영문품목명':'품목명','출고수량':'출고수량합계'})
                        .sort_values('품목명')
                    )
                    st.dataframe(summary_911, use_container_width=True, hide_index=True)

                with st.expander("✅ 재고 충족 SKU (출고 불필요)"):
                    full_items = df_full[['영문품목명','컬러','사이즈','현재고']].rename(columns={'영문품목명':'품목명'})
                    st.dataframe(full_items, use_container_width=True, hide_index=True)

                st.divider()
                tpl_name = template_file_911.name.replace('.xlsx','')
                date_911 = datetime.now().strftime("%Y%m%d")
                st.download_button(
                    "⬇️ 완성된 입고 양식 다운로드",
                    data=buf_911,
                    file_name=f"{tpl_name}_출고수량입력_{date_911}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

            except Exception as e:
                st.error(f"❌ 오류: {e}")
                import traceback
                st.code(traceback.format_exc())
    elif stock_file_911 and not template_file_911:
        st.warning("① 입고 양식 파일도 업로드해주세요.")
    elif template_file_911 and not stock_file_911:
        st.warning("② 매장 재고현황 파일도 업로드해주세요.")
    else:
        st.info("👆 입고 양식과 매장 재고현황 파일을 모두 업로드해주세요.")

# ════════════════════════════════════════════════════════
# TAB 4: 💱 국가별 가격 생성기 (개선 버전)
# ════════════════════════════════════════════════════════

import requests
from datetime import datetime, timedelta

# 전체 국가 목록 (180개국+)
COUNTRY_DATA = {
    'AED': {'korean': '아랍에미리트', 'symbol': 'د.إ', 'region': '중동'},
    'AFN': {'korean': '아프가니스탄', 'symbol': 'Af', 'region': '아시아'},
    'ALL': {'korean': '알바니아', 'symbol': 'L', 'region': '유럽'},
    'AMD': {'korean': '아르메니아', 'symbol': '֏', 'region': '아시아'},
    'ANG': {'korean': '네덜란드령 앤틸레스', 'symbol': 'ƒ', 'region': '카리브'},
    'AOA': {'korean': '앙골라', 'symbol': 'Kz', 'region': '아프리카'},
    'ARS': {'korean': '아르헨티나', 'symbol': '$', 'region': '남미'},
    'AUD': {'korean': '호주', 'symbol': 'A$', 'region': '오세아니아'},
    'AWG': {'korean': '아루바', 'symbol': 'ƒ', 'region': '카리브'},
    'AZN': {'korean': '아제르바이잔', 'symbol': '₼', 'region': '아시아'},
    'BAM': {'korean': '보스니아', 'symbol': 'KM', 'region': '유럽'},
    'BBD': {'korean': '바베이도스', 'symbol': '$', 'region': '카리브'},
    'BDT': {'korean': '방글라데시', 'symbol': '৳', 'region': '아시아'},
    'BGN': {'korean': '불가리아', 'symbol': 'лв', 'region': '유럽'},
    'BHD': {'korean': '바레인', 'symbol': '.د.ب', 'region': '중동'},
    'BIF': {'korean': '부룬디', 'symbol': 'FBu', 'region': '아프리카'},
    'BMD': {'korean': '버뮤다', 'symbol': '$', 'region': '카리브'},
    'BND': {'korean': '브루나이', 'symbol': '$', 'region': '아시아'},
    'BOB': {'korean': '볼리비아', 'symbol': 'Bs.', 'region': '남미'},
    'BRL': {'korean': '브라질', 'symbol': 'R$', 'region': '남미'},
    'BSD': {'korean': '바하마', 'symbol': '$', 'region': '카리브'},
    'BTC': {'korean': '비트코인', 'symbol': '฿', 'region': '암호화폐'},
    'BTN': {'korean': '부탄', 'symbol': 'Nu.', 'region': '아시아'},
    'BWP': {'korean': '보츠와나', 'symbol': 'P', 'region': '아프리카'},
    'BYN': {'korean': '벨라루스', 'symbol': 'Br', 'region': '유럽'},
    'BZD': {'korean': '벨리즈', 'symbol': '$', 'region': '중미'},
    'CAD': {'korean': '캐나다', 'symbol': 'C$', 'region': '북미'},
    'CDF': {'korean': '콩고', 'symbol': 'FC', 'region': '아프리카'},
    'CHE': {'korean': '스위스 (ECU)', 'symbol': 'CHE', 'region': '유럽'},
    'CHF': {'korean': '스위스', 'symbol': 'CHF', 'region': '유럽'},
    'CHW': {'korean': '스위스 (WIR)', 'symbol': 'CHW', 'region': '유럽'},
    'CLF': {'korean': '칠레 (UF)', 'symbol': 'UF', 'region': '남미'},
    'CLP': {'korean': '칠레', 'symbol': '$', 'region': '남미'},
    'CNY': {'korean': '중국', 'symbol': '¥', 'region': '아시아'},
    'COP': {'korean': '콜롬비아', 'symbol': '$', 'region': '남미'},
    'COU': {'korean': '콜롬비아 (UVR)', 'symbol': 'UVR', 'region': '남미'},
    'CRC': {'korean': '코스타리카', 'symbol': '₡', 'region': '중미'},
    'CUC': {'korean': '쿠바 (태환)', 'symbol': '$', 'region': '카리브'},
    'CUP': {'korean': '쿠바', 'symbol': '₱', 'region': '카리브'},
    'CVE': {'korean': '카보베르데', 'symbol': '$', 'region': '아프리카'},
    'CZK': {'korean': '체코', 'symbol': 'Kč', 'region': '유럽'},
    'DJF': {'korean': '지부티', 'symbol': 'Fdj', 'region': '아프리카'},
    'DKK': {'korean': '덴마크', 'symbol': 'kr', 'region': '유럽'},
    'DOP': {'korean': '도미니카공화국', 'symbol': '$', 'region': '카리브'},
    'DZD': {'korean': '알제리', 'symbol': 'د.ج', 'region': '아프리카'},
    'EGP': {'korean': '이집트', 'symbol': '£', 'region': '아프리카'},
    'ERN': {'korean': '에리트레아', 'symbol': 'Nfk', 'region': '아프리카'},
    'ETB': {'korean': '에티오피아', 'symbol': 'Br', 'region': '아프리카'},
    'EUR': {'korean': '유로존', 'symbol': '€', 'region': '유럽'},
    'FJD': {'korean': '피지', 'symbol': '$', 'region': '오세아니아'},
    'FKP': {'korean': '포클랜드', 'symbol': '£', 'region': '남미'},
    'GBP': {'korean': '영국', 'symbol': '£', 'region': '유럽'},
    'GEL': {'korean': '조지아', 'symbol': '₾', 'region': '아시아'},
    'GGP': {'korean': '건지', 'symbol': '£', 'region': '유럽'},
    'GHS': {'korean': '가나', 'symbol': '₵', 'region': '아프리카'},
    'GIP': {'korean': '지브롤터', 'symbol': '£', 'region': '유럽'},
    'GMD': {'korean': '감비아', 'symbol': 'D', 'region': '아프리카'},
    'GNF': {'korean': '기니', 'symbol': 'FG', 'region': '아프리카'},
    'GTQ': {'korean': '과테말라', 'symbol': 'Q', 'region': '중미'},
    'GYD': {'korean': '가이아나', 'symbol': '$', 'region': '남미'},
    'HKD': {'korean': '홍콩', 'symbol': '$', 'region': '아시아'},
    'HNL': {'korean': '온두라스', 'symbol': 'L', 'region': '중미'},
    'HRK': {'korean': '크로아티아', 'symbol': 'kn', 'region': '유럽'},
    'HTG': {'korean': '아이티', 'symbol': 'G', 'region': '카리브'},
    'HUF': {'korean': '헝가리', 'symbol': 'Ft', 'region': '유럽'},
    'IDR': {'korean': '인도네시아', 'symbol': 'Rp', 'region': '아시아'},
    'ILS': {'korean': '이스라엘', 'symbol': '₪', 'region': '중동'},
    'IMP': {'korean': '맨섬', 'symbol': '£', 'region': '유럽'},
    'INR': {'korean': '인도', 'symbol': '₹', 'region': '아시아'},
    'IQD': {'korean': '이라크', 'symbol': 'ع.د', 'region': '중동'},
    'IRR': {'korean': '이란', 'symbol': '﷼', 'region': '중동'},
    'ISK': {'korean': '아이슬란드', 'symbol': 'kr', 'region': '유럽'},
    'JEP': {'korean': '저지', 'symbol': '£', 'region': '유럽'},
    'JMD': {'korean': '자메이카', 'symbol': '$', 'region': '카리브'},
    'JOD': {'korean': '요르단', 'symbol': 'د.ا', 'region': '중동'},
    'JPY': {'korean': '일본', 'symbol': '¥', 'region': '아시아'},
    'KES': {'korean': '케냐', 'symbol': 'Sh', 'region': '아프리카'},
    'KGS': {'korean': '키르기스스탄', 'symbol': 'с', 'region': '아시아'},
    'KHR': {'korean': '캄보디아', 'symbol': '៛', 'region': '아시아'},
    'KMF': {'korean': '코모로', 'symbol': 'CF', 'region': '아프리카'},
    'KPW': {'korean': '북한', 'symbol': '₩', 'region': '아시아'},
    'KRW': {'korean': '대한민국', 'symbol': '₩', 'region': '아시아'},
    'KWD': {'korean': '쿠웨이트', 'symbol': 'د.ك', 'region': '중동'},
    'KYD': {'korean': '케이맨 제도', 'symbol': '$', 'region': '카리브'},
    'KZT': {'korean': '카자흐스탄', 'symbol': '₸', 'region': '아시아'},
    'LAK': {'korean': '라오스', 'symbol': '₭', 'region': '아시아'},
    'LBP': {'korean': '레바논', 'symbol': '£', 'region': '중동'},
    'LKR': {'korean': '스리랑카', 'symbol': 'Rs', 'region': '아시아'},
    'LRD': {'korean': '라이베리아', 'symbol': '$', 'region': '아프리카'},
    'LSL': {'korean': '레소토', 'symbol': 'L', 'region': '아프리카'},
    'LYD': {'korean': '리비아', 'symbol': 'ل.د', 'region': '아프리카'},
    'MAD': {'korean': '모로코', 'symbol': 'د.م.', 'region': '아프리카'},
    'MDL': {'korean': '몰도바', 'symbol': 'L', 'region': '유럽'},
    'MGA': {'korean': '마다가스카르', 'symbol': 'Ar', 'region': '아프리카'},
    'MKD': {'korean': '북마케도니아', 'symbol': 'ден', 'region': '유럽'},
    'MMK': {'korean': '미얀마', 'symbol': 'K', 'region': '아시아'},
    'MNT': {'korean': '몽골', 'symbol': '₮', 'region': '아시아'},
    'MOP': {'korean': '마카오', 'symbol': 'P', 'region': '아시아'},
    'MRU': {'korean': '모리타니', 'symbol': 'UM', 'region': '아프리카'},
    'MUR': {'korean': '모리셔스', 'symbol': '₨', 'region': '아프리카'},
    'MVR': {'korean': '몰디브', 'symbol': '.ރ.', 'region': '아시아'},
    'MWK': {'korean': '말라위', 'symbol': 'K', 'region': '아프리카'},
    'MXN': {'korean': '멕시코', 'symbol': '$', 'region': '북미'},
    'MYR': {'korean': '말레이시아', 'symbol': 'RM', 'region': '아시아'},
    'MZN': {'korean': '모잠비크', 'symbol': 'MT', 'region': '아프리카'},
    'NAD': {'korean': '나미비아', 'symbol': '$', 'region': '아프리카'},
    'NGN': {'korean': '나이지리아', 'symbol': '₦', 'region': '아프리카'},
    'NIO': {'korean': '니카라과', 'symbol': 'C$', 'region': '중미'},
    'NOK': {'korean': '노르웨이', 'symbol': 'kr', 'region': '유럽'},
    'NPR': {'korean': '네팔', 'symbol': '₨', 'region': '아시아'},
    'NZD': {'korean': '뉴질랜드', 'symbol': 'NZ$', 'region': '오세아니아'},
    'OMR': {'korean': '오만', 'symbol': 'ر.ع.', 'region': '중동'},
    'PAB': {'korean': '파나마', 'symbol': 'B/.', 'region': '중미'},
    'PEN': {'korean': '페루', 'symbol': 'S/', 'region': '남미'},
    'PGK': {'korean': '파푸아뉴기니', 'symbol': 'K', 'region': '오세아니아'},
    'PHP': {'korean': '필리핀', 'symbol': '₱', 'region': '아시아'},
    'PKR': {'korean': '파키스탄', 'symbol': '₨', 'region': '아시아'},
    'PLN': {'korean': '폴란드', 'symbol': 'zł', 'region': '유럽'},
    'PYG': {'korean': '파라과이', 'symbol': '₲', 'region': '남미'},
    'QAR': {'korean': '카타르', 'symbol': 'ر.ق', 'region': '중동'},
    'RON': {'korean': '루마니아', 'symbol': 'lei', 'region': '유럽'},
    'RSD': {'korean': '세르비아', 'symbol': 'дин.', 'region': '유럽'},
    'RUB': {'korean': '러시아', 'symbol': '₽', 'region': '유럽'},
    'RWF': {'korean': '르완다', 'symbol': 'FRw', 'region': '아프리카'},
    'SAR': {'korean': '사우디아라비아', 'symbol': 'ر.س', 'region': '중동'},
    'SBD': {'korean': '솔로몬제도', 'symbol': '$', 'region': '오세아니아'},
    'SCR': {'korean': '세이셀', 'symbol': '₨', 'region': '아프리카'},
    'SDG': {'korean': '수단', 'symbol': '£', 'region': '아프리카'},
    'SEK': {'korean': '스웨덴', 'symbol': 'kr', 'region': '유럽'},
    'SGD': {'korean': '싱가포르', 'symbol': '$', 'region': '아시아'},
    'SHP': {'korean': '세인트헬레나', 'symbol': '£', 'region': '아프리카'},
    'SLL': {'korean': '시에라리온', 'symbol': 'Le', 'region': '아프리카'},
    'SOS': {'korean': '소말리아', 'symbol': 'Sh', 'region': '아프리카'},
    'SRD': {'korean': '수리남', 'symbol': '$', 'region': '남미'},
    'SSP': {'korean': '남수단', 'symbol': '£', 'region': '아프리카'},
    'STN': {'korean': '상투메프린시페', 'symbol': 'Db', 'region': '아프리카'},
    'SYP': {'korean': '시리아', 'symbol': '£', 'region': '중동'},
    'SZL': {'korean': '에스와티니', 'symbol': 'E', 'region': '아프리카'},
    'THB': {'korean': '태국', 'symbol': '฿', 'region': '아시아'},
    'TJS': {'korean': '타지키스탄', 'symbol': 'ЅМ', 'region': '아시아'},
    'TMT': {'korean': '투르크메니스탄', 'symbol': 'm', 'region': '아시아'},
    'TND': {'korean': '튀니지', 'symbol': 'د.ت', 'region': '아프리카'},
    'TOP': {'korean': '통가', 'symbol': 'T$', 'region': '오세아니아'},
    'TRY': {'korean': '터키', 'symbol': '₺', 'region': '중동'},
    'TTD': {'korean': '트리니다드토바고', 'symbol': '$', 'region': '카리브'},
    'TWD': {'korean': '대만', 'symbol': 'NT$', 'region': '아시아'},
    'TZS': {'korean': '탄자니아', 'symbol': 'Sh', 'region': '아프리카'},
    'UAH': {'korean': '우크라이나', 'symbol': '₴', 'region': '유럽'},
    'UGX': {'korean': '우간다', 'symbol': 'Sh', 'region': '아프리카'},
    'USD': {'korean': '미국', 'symbol': '$', 'region': '북미'},
    'UYU': {'korean': '우루과이', 'symbol': '$', 'region': '남미'},
    'UZS': {'korean': '우즈베키스탄', 'symbol': 'so\'m', 'region': '아시아'},
    'VEF': {'korean': '베네수엘라 (구)', 'symbol': 'Bs.', 'region': '남미'},
    'VES': {'korean': '베네수엘라', 'symbol': 'Bs.S', 'region': '남미'},
    'VND': {'korean': '베트남', 'symbol': '₫', 'region': '아시아'},
    'VUV': {'korean': '바누아투', 'symbol': 'VT', 'region': '오세아니아'},
    'WST': {'korean': '사모아', 'symbol': 'T', 'region': '오세아니아'},
    'XAF': {'korean': '중앙아프리카CFA', 'symbol': 'FCFA', 'region': '아프리카'},
    'XCD': {'korean': '동카리브', 'symbol': '$', 'region': '카리브'},
    'XOF': {'korean': '서아프리카CFA', 'symbol': 'CFA', 'region': '아프리카'},
    'XPF': {'korean': '태평양CFP', 'symbol': '₣', 'region': '오세아니아'},
    'YER': {'korean': '예멘', 'symbol': '﷼', 'region': '중동'},
    'ZAR': {'korean': '남아프리카', 'symbol': 'R', 'region': '아프리카'},
    'ZMW': {'korean': '잠비아', 'symbol': 'K', 'region': '아프리카'},
    'ZWL': {'korean': '짐바브웨', 'symbol': 'Z$', 'region': '아프리카'},
}

# 기본 조정비율 (일부만 지정, 나머지는 0%)
DEFAULT_ADJUSTMENTS = {
    'USD': 68.3, 'EUR': 0, 'GBP': 54.0, 'JPY': 39.7, 'CNY': 39.7,
    'AED': 29.8, 'CAD': 39.7, 'AUD': 30.9, 'NZD': 38.6, 'CHF': 36.4,
    'SEK': 60.6, 'NOK': 56.2, 'DKK': 60.6, 'HUF': 61.7, 'PLN': 57.3, 'CZK': 55.1,
}

@st.cache_data(ttl=3600)  # 1시간 캐시
def get_exchange_rates_from_api():
    """exchangerate-api에서 실시간 환율 조회"""
    try:
        url = 'https://api.exchangerate-api.com/v4/latest/KRW'
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        rates = response.json()['rates']
        st.success("✅ 환율 자동 조회 완료")
        return rates
    except Exception as e:
        st.warning(f"⚠️ 환율 조회 실패: {str(e)[:50]}")
        return None

def make_country_pricing_sheet(product_file, country_list, exchange_rates):
    """프로덕트 시트에서 기본가격을 읽고 국가별 가격 시트 생성"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    prod_bytes = product_file.read() if hasattr(product_file, 'read') else open(product_file, 'rb').read()
    wb_prod = load_workbook(io.BytesIO(prod_bytes), data_only=True)
    ws_prod = wb_prod.active
    
    products = []
    for row in range(4, ws_prod.max_row + 1):
        cat = ws_prod.cell(row, 1).value
        prod_name = ws_prod.cell(row, 2).value
        style_no = ws_prod.cell(row, 3).value
        color = ws_prod.cell(row, 4).value
        price_krw = ws_prod.cell(row, 13).value
        
        if prod_name and style_no and price_krw:
            try:
                price_krw = float(price_krw)
                products.append({
                    'category': cat,
                    'product_name': prod_name,
                    'style_no': style_no,
                    'color': color,
                    'price_krw': price_krw,
                })
            except:
                pass
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Country Pricing'
    
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = ['Category', 'Product Name', 'Style No.', 'Color', 'Base Price (KRW)']
    headers += [f"{c['code']} ({c['symbol']})" for c in country_list]
    
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006FC0')
        cell.alignment = center_align
        cell.border = border_all
    
    ws_out.column_dimensions['A'].width = 15
    ws_out.column_dimensions['B'].width = 40
    ws_out.column_dimensions['C'].width = 16
    ws_out.column_dimensions['D'].width = 20
    ws_out.column_dimensions['E'].width = 18
    
    for col in range(6, 6 + len(country_list)):
        ws_out.column_dimensions[get_column_letter(col)].width = 16
    
    for row_idx, prod in enumerate(products, 2):
        cell = ws_out.cell(row=row_idx, column=1, value=prod['category'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=2, value=prod['product_name'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=3, value=prod['style_no'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=4, value=prod['color'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=5, value=int(prod['price_krw']))
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0'
        
        for col_idx, country in enumerate(country_list, 6):
            code = country['code']
            adjustment = country['adjustment']
            exchange = exchange_rates.get(code, 1.0)
            
            price_local = (prod['price_krw'] / exchange) * (1 + adjustment)
            
            cell = ws_out.cell(row=row_idx, column=col_idx, value=price_local)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = center_align
            cell.border = border_all
            cell.number_format = '#,##0.00'
    
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


def make_single_country_sheet(product_file, country, exchange_rate):
    """단일 국가별 가격 시트 생성"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    prod_bytes = product_file.read() if hasattr(product_file, 'read') else open(product_file, 'rb').read()
    wb_prod = load_workbook(io.BytesIO(prod_bytes), data_only=True)
    ws_prod = wb_prod.active
    
    products = []
    for row in range(4, ws_prod.max_row + 1):
        cat = ws_prod.cell(row, 1).value
        prod_name = ws_prod.cell(row, 2).value
        style_no = ws_prod.cell(row, 3).value
        color = ws_prod.cell(row, 4).value
        price_krw = ws_prod.cell(row, 13).value
        
        if prod_name and style_no and price_krw:
            try:
                price_krw = float(price_krw)
                products.append({
                    'category': cat,
                    'product_name': prod_name,
                    'style_no': style_no,
                    'color': color,
                    'price_krw': price_krw,
                })
            except:
                pass
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = country['code']
    
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = ['Category', 'Product Name', 'Style No.', 'Color', 'Base Price (KRW)', f"{country['code']} ({country['symbol']})"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006FC0')
        cell.alignment = center_align
        cell.border = border_all
    
    ws_out.column_dimensions['A'].width = 15
    ws_out.column_dimensions['B'].width = 40
    ws_out.column_dimensions['C'].width = 16
    ws_out.column_dimensions['D'].width = 20
    ws_out.column_dimensions['E'].width = 18
    ws_out.column_dimensions['F'].width = 18
    
    for row_idx, prod in enumerate(products, 2):
        cell = ws_out.cell(row=row_idx, column=1, value=prod['category'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=2, value=prod['product_name'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=3, value=prod['style_no'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=4, value=prod['color'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_align
        cell.border = border_all
        
        cell = ws_out.cell(row=row_idx, column=5, value=int(prod['price_krw']))
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0'
        
        adjustment = country['adjustment']
        price_local = (prod['price_krw'] / exchange_rate) * (1 + adjustment)
        
        cell = ws_out.cell(row=row_idx, column=6, value=price_local)
        cell.font = Font(name='Arial', size=9, bold=True, color='006FC0')
        cell.alignment = center_align
        cell.border = border_all
        cell.number_format = '#,##0.00'
    
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


# ── TAB 4: 국가별 가격 생성기
with tab4:
    st.caption("🌍 프로덕트 시트 + 실시간 환율 + 조정비율 → 국가별 가격 시트 자동 생성")
    st.divider()
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("📁 파일 업로드")
        pricing_product_file = st.file_uploader(
            "프로덕트 시트 Excel (.xlsx)",
            type=['xlsx'],
            key="pricing_product"
        )
        st.caption("📌 예: _SPECIALGUEST____NEO_DAWN___2026SS____PRODUCT_LIST.xlsx")
    
    with col2:
        st.subheader("🌐 환율 연동")
        if st.button("🔄 실시간 환율 조회", use_container_width=True, type="primary"):
            api_rates = get_exchange_rates_from_api()
            if api_rates:
                st.session_state['api_rates'] = api_rates
                st.info(f"✅ {len(api_rates)}개 국가 환율 로드됨")
    
    st.divider()
    
    if pricing_product_file:
        st.subheader("🔍 국가 검색 및 선택")
        
        # 검색 박스
        search_query = st.text_input(
            "국가명 또는 통화 코드로 검색 (예: 미국, USD, 일본, JPY)",
            placeholder="검색...",
            key="country_search"
        )
        
        # 필터링
        filtered_countries = {}
        for code, data in COUNTRY_DATA.items():
            korean_name = data['korean']
            if search_query.lower() in code.lower() or \
               search_query in korean_name or \
               code.lower() in search_query.lower():
                filtered_countries[code] = data
        
        # 검색 결과가 없으면 전체 표시
        if not search_query:
            filtered_countries = COUNTRY_DATA
        
        st.caption(f"📊 검색 결과: {len(filtered_countries)}개국")
        
        # 체크박스로 국가 선택
        selected_codes = []
        cols = st.columns(4)
        col_idx = 0
        
        for code in sorted(filtered_countries.keys()):
            data = filtered_countries[code]
            with cols[col_idx % 4]:
                if st.checkbox(
                    f"{code} - {data['korean']} ({data['symbol']})",
                    key=f"country_{code}"
                ):
                    selected_codes.append(code)
            col_idx += 1
        
        if selected_codes:
            st.success(f"✅ {len(selected_codes)}개 국가 선택됨: {', '.join(selected_codes)}")
        else:
            st.warning("⚠️ 최소 1개 국가를 선택해주세요.")
        
        st.divider()
        
        if selected_codes:
            st.subheader("💱 환율 및 조정비율 설정")
            
            # API 환율 로드
            api_rates = st.session_state.get('api_rates')
            
            rates = {}
            adjustments = {}
            
            cols = st.columns(len(selected_codes))
            
            for idx, code in enumerate(selected_codes):
                with cols[idx]:
                    data = COUNTRY_DATA[code]
                    st.write(f"**{code}** - {data['korean']}")
                    
                    # 기본값: API 환율 또는 수동입력
                    default_rate = 1.0
                    if api_rates and code in api_rates:
                        default_rate = api_rates[code]
                        st.caption(f"📊 API 환율: {default_rate:.4f}")
                    
                    rates[code] = st.number_input(
                        f"환율 (KRW/1{code})",
                        value=float(default_rate),
                        step=0.1,
                        key=f"rate_{code}",
                        label_visibility="collapsed",
                        help=f"1 {code} = ? KRW"
                    )
                    
                    default_adj = DEFAULT_ADJUSTMENTS.get(code, 0)
                    adj_pct = st.number_input(
                        f"조정 %",
                        value=int(default_adj),
                        step=1,
                        key=f"adj_{code}",
                        label_visibility="collapsed",
                        help="조정 비율 (%)"
                    )
                    adjustments[code] = adj_pct / 100
            
            st.divider()
            
            col_btn1, col_btn2 = st.columns([1, 1])
            
            with col_btn1:
                if st.button("📊 통합 시트 생성 (모든 국가)", use_container_width=True, type="primary"):
                    with st.spinner("국가별 가격 계산 중..."):
                        try:
                            country_data = []
                            for code in selected_codes:
                                data = COUNTRY_DATA[code]
                                country_data.append({
                                    'code': code,
                                    'name': data['korean'],
                                    'symbol': data['symbol'],
                                    'adjustment': adjustments.get(code, 0),
                                })
                            
                            buf = make_country_pricing_sheet(
                                pricing_product_file,
                                country_data,
                                rates
                            )
                            
                            st.success("✅ 통합 시트 생성 완료!")
                            
                            date_str = datetime.now().strftime("%Y%m%d")
                            st.download_button(
                                "⬇️ 통합 시트 다운로드",
                                data=buf,
                                file_name=f"Country_Pricing_All_{date_str}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary"
                            )
                            
                        except Exception as e:
                            st.error(f"❌ 오류: {e}")
                            import traceback
                            st.code(traceback.format_exc())
            
            with col_btn2:
                if st.button("🌍 국가별 개별 시트 생성", use_container_width=True, type="secondary"):
                    with st.spinner("국가별 시트 생성 중..."):
                        try:
                            date_str = datetime.now().strftime("%Y%m%d")
                            
                            individual_files = []
                            for code in selected_codes:
                                data = COUNTRY_DATA[code]
                                country_info = {
                                    'code': code,
                                    'korean': data['korean'],
                                    'symbol': data['symbol'],
                                    'adjustment': adjustments.get(code, 0),
                                }
                                
                                buf = make_single_country_sheet(
                                    pricing_product_file,
                                    country_info,
                                    rates[code]
                                )
                                
                                individual_files.append({
                                    'code': code,
                                    'buf': buf,
                                    'filename': f"Pricing_{code}_{date_str}.xlsx"
                                })
                            
                            st.success(f"✅ {len(individual_files)}개 국가 시트 생성 완료!")
                            
                            st.subheader("📥 국가별 다운로드")
                            cols = st.columns(len(individual_files))
                            
                            for idx, file_info in enumerate(individual_files):
                                with cols[idx]:
                                    st.download_button(
                                        f"⬇️ {file_info['code']}",
                                        data=file_info['buf'],
                                        file_name=file_info['filename'],
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )
                            
                        except Exception as e:
                            st.error(f"❌ 오류: {e}")
                            import traceback
                            st.code(traceback.format_exc())
    else:
        st.info("👆 프로덕트 시트 파일을 업로드해주세요.")

st.divider()
st.caption("💡 팁: 🔄 실시간 환율 조회 버튼으로 API에서 환율을 불러올 수 있습니다. 원하는 값으로 수정 가능합니다!")
